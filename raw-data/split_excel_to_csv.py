import os, csv
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, 'Endfield伤害期望计算器v0.12.xlsx')
OUT_DIR = BASE_DIR

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    csv_name = sheet_name.replace('/', '_').replace('\\', '_') + '.csv'
    csv_path = os.path.join(OUT_DIR, csv_name)

    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        cleaned = []
        for cell in row:
            if cell is None:
                cleaned.append('')
            else:
                cleaned.append(str(cell) if not isinstance(cell, (int, float)) else cell)
        rows.append(cleaned)

    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    print(f'Exported: {csv_name} ({len(rows)} rows)')

print('Done.')
